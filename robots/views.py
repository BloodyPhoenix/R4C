import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from django.views import View
from .models import Robot


class DownloadWeeklyReportView(View):

    @staticmethod
    def get_last_week_production():
        # We are counting all the data from the day seven days ago, not just last 168 hours
        today_date = datetime.date.today()
        time = datetime.datetime.strptime('00:00:00', '%H:%M:%S').time()
        today_full_date = datetime.datetime.combine(today_date, time)
        week_ago = today_full_date - datetime.timedelta(days=7)
        queryset = Robot.objects.filter(created__gte=week_ago)
        return queryset

    @staticmethod
    def format_queryset_data(queryset):
        # Sorting and converting queryset data before adding it into file
        formatted_data = dict()
        for robot in queryset:
            model = robot.model
            version = robot.version
            if model in formatted_data:
                if version in formatted_data[model]:
                    formatted_data[model][version] += 1
                else:
                    formatted_data[model][version] = 1
            else:
                formatted_data[model] = {version: 1}
        return formatted_data

    @staticmethod
    def get_excel_file(data: dict, filename=None, response=None):
        # I left Workbook object generation for simplifying testing.
        # If we are not returning Response object we can check correct data in file via unittests.
        # filename argument os also left to simplify testing
        if filename is None:
            filename = 'weekly_production_report ' + str(datetime.date.today()) + '.xlsx'
        report = Workbook()
        report.remove(report.active)
        headers = ["Модель", "Версия", "Количество за неделю"]
        if len(data) == 0:
            report.create_sheet("Нет данных за прошедшую неделю")
            report.active = report["Нет данных за прошедшую неделю"]
            worksheet = report.active
            worksheet.append(headers)
        else:
            for model, version in data.items():
                report.create_sheet(model)
                report.active = report[model]
                worksheet = report.active
                worksheet.append(headers)
                for version_name, quantity in version.items():
                    row = [model, version_name, quantity]
                    worksheet.append(row)
        if response:
            report.save(response)
            return response
        else:
            report.save(filename)
            return report

    def get(self, request, *args, **kwargs):
        queryset = self.get_last_week_production()
        formatted_data = self.format_queryset_data(queryset)
        filename = 'weekly_production_report ' + str(datetime.date.today()) + '.xlsx'
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        report = self.get_excel_file(formatted_data, filename, response)
        return report

